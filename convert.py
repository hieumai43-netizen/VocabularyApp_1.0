import openpyxl
import json
import os

EXCEL_FILE = "vocabulary_database_template.xlsx"
OUTPUT_FILE = "data.js"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(BASE_DIR, EXCEL_FILE)
output_path = os.path.join(BASE_DIR, OUTPUT_FILE)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

headers = []
for cell in ws[1]:
    headers.append(str(cell.value).strip())

data = []

for row in ws.iter_rows(min_row=2, values_only=True):
    item = dict(zip(headers, row))

    if item.get("ID") is None:
        continue

    vocab_id = str(item.get("ID")).strip()

    word = {
        "id": vocab_id,
        "category": item.get("Category") or "",
        "vn": item.get("Vietnamese") or "",
        "jp": item.get("Japanese") or "",
        "kana": item.get("Kana") or "",
        "en": item.get("English") or "",
        "ipa": item.get("IPA") or "",
        "cn": item.get("Chinese") or "",
        "pinyin": item.get("Pinyin") or "",
        "ko": item.get("Korean") or "",
        "koread": item.get("KoreanReading") or "",
        "example_vi": item.get("Example_VI") or "",
        "example_ja": item.get("Example_JA") or "",
        "example_en": item.get("Example_EN") or "",
        "example_cn": item.get("Example_CN") or "",
        "example_ko": item.get("Example_KO") or "",
        "note": item.get("Note") or "",

        # Tự sinh đường dẫn MP3 theo ID
        "audio_word": f"audio/{vocab_id}_word.mp3",
        "audio_example": f"audio/{vocab_id}_example.mp3"
    }

    data.append(word)

with open(output_path, "w", encoding="utf-8") as f:
    f.write("const vocabularyData = ")
    json.dump(data, f, ensure_ascii=False, indent=2)
    f.write(";")

print("✅ Convert thành công!")
print(f"📘 Đã đọc: {EXCEL_FILE}")
print(f"📄 Đã tạo: {OUTPUT_FILE}")
print(f"🔊 Audio sẽ tự hiểu dạng: audio/ID_word.mp3 và audio/ID_example.mp3")
print(f"📦 Tổng số từ: {len(data)}")