from gtts import gTTS
from openpyxl import load_workbook
import os
import json

EXCEL_FILE = "vocabulary_database_template.xlsx"
AUDIO_DIR = "audio"
JSON_FILE = "data.json"

os.makedirs(AUDIO_DIR, exist_ok=True)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

headers = [cell.value for cell in ws[1]]

def get(row, name):
    if name not in headers:
        return ""
    value = row[headers.index(name)]
    return "" if value is None else str(value).strip()

data = []
seen_ids = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    raw_id = get(row, "ID")

    if raw_id == "":
        continue

    word_id = str(int(float(raw_id))).zfill(3)

    if word_id in seen_ids:
        print(f"LỖI: Trùng ID {word_id}")
        continue

    seen_ids.add(word_id)

    item = {
        "id": word_id,
        "category": get(row, "Category"),
        "vietnamese": get(row, "Vietnamese"),

        "japanese": get(row, "Japanese"),
        "kana": get(row, "Kana"),

        "english": get(row, "English"),
        "ipa": get(row, "IPA"),

        "chinese": get(row, "Chinese"),
        "pinyin": get(row, "Pinyin"),

        "korean": get(row, "Korean"),
        "koreanReading": get(row, "KoreanReading"),

        "example_vi": get(row, "Example_VI"),
        "example_ja": get(row, "Example_JA"),
        "example_en": get(row, "Example_EN"),
        "example_cn": get(row, "Example_CN"),
        "example_ko": get(row, "Example_KO"),

        "audio": {
            "jp": f"audio/{word_id}/jp.mp3",
            "en": f"audio/{word_id}/en.mp3",
            "cn": f"audio/{word_id}/cn.mp3",
            "ko": f"audio/{word_id}/ko.mp3",
            "ex_jp": f"audio/{word_id}/ex_jp.mp3",
            "ex_en": f"audio/{word_id}/ex_en.mp3",
            "ex_cn": f"audio/{word_id}/ex_cn.mp3",
            "ex_ko": f"audio/{word_id}/ex_ko.mp3",
        }
    }

    data.append(item)

    word_dir = os.path.join(AUDIO_DIR, word_id)
    os.makedirs(word_dir, exist_ok=True)

    mp3_list = [
        (item["japanese"], "ja", "jp.mp3"),
        (item["english"], "en", "en.mp3"),
        (item["chinese"], "zh-CN", "cn.mp3"),
        (item["korean"], "ko", "ko.mp3"),

        (item["example_ja"], "ja", "ex_jp.mp3"),
        (item["example_en"], "en", "ex_en.mp3"),
        (item["example_cn"], "zh-CN", "ex_cn.mp3"),
        (item["example_ko"], "ko", "ex_ko.mp3"),
    ]

    for text, lang, filename in mp3_list:
        if text == "":
            continue

        file_path = os.path.join(word_dir, filename)

        if os.path.exists(file_path):
            print(f"Bỏ qua MP3 đã có: {file_path}")
            continue

        tts = gTTS(text=text, lang=lang)
        tts.save(file_path)
        print(f"Đã tạo MP3: {file_path}")

with open(JSON_FILE, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("--------------------------------")
print(f"Hoàn thành!")
print(f"Đã tạo/cập nhật: {JSON_FILE}")
print(f"Tổng số từ: {len(data)}")
print("--------------------------------")