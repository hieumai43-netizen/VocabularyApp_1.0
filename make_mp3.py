from gtts import gTTS
from openpyxl import load_workbook
import os

EXCEL_FILE = "vocabulary_database_template.xlsx"
AUDIO_DIR = "audio"

os.makedirs(AUDIO_DIR, exist_ok=True)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

headers = [cell.value for cell in ws[1]]

def get(row, name):
    if name not in headers:
        return None
    return row[headers.index(name)]

for row in ws.iter_rows(min_row=2, values_only=True):
    word_id = get(row, "ID")

    if word_id is None:
        continue

    word_id = str(int(word_id)).zfill(3)

    word_dir = os.path.join(AUDIO_DIR, word_id)
    os.makedirs(word_dir, exist_ok=True)

    mp3_list = [
        (get(row, "Japanese"), "ja", "jp.mp3"),
        (get(row, "English"), "en", "en.mp3"),
        (get(row, "Chinese"), "zh-CN", "cn.mp3"),
        (get(row, "Korean"), "ko", "ko.mp3"),

        (get(row, "Example_JA"), "ja", "ex_jp.mp3"),
        (get(row, "Example_EN"), "en", "ex_en.mp3"),
        (get(row, "Example_CN"), "zh-CN", "ex_cn.mp3"),
        (get(row, "Example_KO"), "ko", "ex_ko.mp3"),
    ]

    for text, lang, filename in mp3_list:
        file_path = os.path.join(word_dir, filename)

        if text is None or str(text).strip() == "":
            continue

        if os.path.exists(file_path):
            print(f"Bỏ qua file đã có: {file_path}")
            continue

        tts = gTTS(text=str(text), lang=lang)
        tts.save(file_path)

        print(f"Đã tạo: {file_path}")

print("Hoàn thành tạo MP3!")